
import io
from datetime import datetime
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Yusen Monthly Tracking Report", layout="wide")

st.title("Yusen Monthly Tracking Report Generator")

st.markdown(
    "Upload the **Data Availability Trend by Selected Dimensions** Excel export. "
    "This app will generate a monthly tenant report in the exact Yusen layout "
    "(months as merged headers; sub-columns for Volume Created, Volume Tracked, "
    "Volume Not Tracked, and Tracked %)."
)

# --- Tenants that must always appear ---
REQUIRED_TENANTS = [
    "Yusen Logistics Benelux B.V.",
    "Yusen Logistics Czech s.r.o.",
    "Yusen Logistics France S.A.S.",
    "Yusen Logistics Poland Sp. z.o.o.",
    "Yusen Logistics Slovakia",
    "Yusen Logistics Germany",
    "Yusen Logistics Hungary",
]

# --- Helpers ---
def to_bool(x):
    if isinstance(x, bool):
        return x
    if pd.isna(x):
        return False
    s = str(x).strip().lower()
    if s in ("true", "1", "yes", "y", "t"):
        return True
    if s in ("false", "0", "no", "n", "f"):
        return False
    return False

def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    # Normalize columns
    if "Tenant Name" not in df.columns:
        raise ValueError("Missing column: 'Tenant Name'")
    if "Tracked" not in df.columns:
        raise ValueError("Missing column: 'Tracked'")
    # The user confirmed the date is in 'Period Date'
    if "Period Date" not in df.columns:
        # fallback if needed
        alt_cols = [c for c in df.columns if "date" in c.lower()]
        raise ValueError("Missing column: 'Period Date' (found alternatives: %s)" % alt_cols)

    # Clean + parse
    df = df.copy()
    df["Tenant Name"] = df["Tenant Name"].fillna("Unknown").astype(str).str.strip()
    df["Tracked"] = df["Tracked"].map(to_bool)
    df["Period Date"] = pd.to_datetime(df["Period Date"], errors="coerce")
    df = df.dropna(subset=["Period Date"])
    df["YearMonth"] = df["Period Date"].dt.to_period("M").astype(str)

    # Ensure required tenants exist even if 0 rows
    existing_tenants = set(df["Tenant Name"].unique())
    missing_required = [t for t in REQUIRED_TENANTS if t not in existing_tenants]
    # Create zero-rows for missing tenants using the months present in df (or at least current month)
    months = sorted(df["YearMonth"].unique().tolist())
    if not months:
        months = [datetime.now().strftime("%Y-%m")]

    if missing_required:
        add_rows = []
        for t in missing_required:
            for m in months:
                add_rows.append({"Tenant Name": t, "Tracked": False, "Period Date": pd.NaT, "YearMonth": m})
        df = pd.concat([df, pd.DataFrame(add_rows)], ignore_index=True)

    # Aggregate
    summary = (
        df.groupby(["Tenant Name", "YearMonth"])
          .agg(
              Volume_Created=("Tracked", "count"),
              Volume_Tracked=("Tracked", lambda x: np.sum(x.astype(bool))),
              Volume_Not_Tracked=("Tracked", lambda x: np.sum(~x.astype(bool))),
          )
          .reset_index()
    )
    summary["Tracked_Percentage"] = np.where(
        summary["Volume_Created"] > 0,
        summary["Volume_Tracked"] / summary["Volume_Created"],
        0.0
    )
    return summary

def to_excel_report(summary: pd.DataFrame, tenants_first: list[str] | None = None) -> bytes:
    # Prepare pivot
    metrics = ["Volume_Created", "Volume_Tracked", "Volume_Not_Tracked", "Tracked_Percentage"]
    pivot = summary.pivot_table(
        index="Tenant Name",
        columns="YearMonth",
        values=metrics,
        fill_value=0,
        aggfunc="first",
    )
    months = sorted(summary["YearMonth"].unique().tolist())

    # Order tenants: REQUIRED_TENANTS first (in given order), then the rest alphabetically
    all_tenants = list(pivot.index.unique())
    remaining = sorted([t for t in all_tenants if t not in REQUIRED_TENANTS])
    ordered_tenants = REQUIRED_TENANTS + remaining
    # Reindex pivot to ordered tenants (ignore missing)
    valid_ordered = [t for t in ordered_tenants if t in all_tenants]
    pivot = pivot.reindex(index=valid_ordered)

    # Write with openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # Header
    ws.cell(row=1, column=1, value="Tenant Name")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws["A1"].font = header_font
    ws["A1"].alignment = center
    ws["A1"].border = border_all

    start_col = 2
    col_pointer = start_col
    for month in months:
        ws.merge_cells(start_row=1, start_column=col_pointer, end_row=1, end_column=col_pointer + len(metrics) - 1)
        ws.cell(row=1, column=col_pointer, value=month).font = header_font
        ws.cell(row=1, column=col_pointer).alignment = center

        for i, m in enumerate(metrics):
            label = m.replace("_", " ")
            ws.cell(row=2, column=col_pointer + i, value=label).font = header_font
            ws.cell(row=2, column=col_pointer + i).alignment = center
        col_pointer += len(metrics)

    # Data rows
    row_pointer = 3
    for tenant in pivot.index:
        ws.cell(row=row_pointer, column=1, value=tenant)
        ws.cell(row=row_pointer, column=1).alignment = left
        ws.cell(row=row_pointer, column=1).border = border_all

        col_pointer = start_col
        for month in months:
            for m in metrics:
                try:
                    value = pivot.loc[tenant, (m, month)]
                except KeyError:
                    value = 0
                cell = ws.cell(row=row_pointer, column=col_pointer, value=value)
                if m == "Tracked_Percentage":
                    cell.number_format = "0.00%"
                    cell.alignment = right
                else:
                    try:
                        cell.value = int(value)
                    except Exception:
                        pass
                    cell.number_format = "0"
                    cell.alignment = right
                cell.border = border_all
                col_pointer += 1
        row_pointer += 1

    # Borders & alignments for headers
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        ws.cell(row=1, column=c).border = border_all
        ws.cell(row=1, column=c).alignment = center
        ws.cell(row=2, column=c).border = border_all
        ws.cell(row=2, column=c).alignment = center

    # Column widths
    ws.column_dimensions["A"].width = 36
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # Freeze panes
    ws.freeze_panes = "B3"

    # Save to bytes
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

# --- UI ---
uploaded = st.file_uploader("Upload the main Excel file (.xlsx)", type=["xlsx"])

if uploaded is not None:
    try:
        df = pd.read_excel(uploaded)
        st.success("File loaded successfully.")
        # Optional quick preview
        with st.expander("Preview first 20 rows"):
            st.dataframe(df.head(20))
        # Build summary
        summary = build_summary(df)

        # Show a small preview of the aggregated data
        st.subheader("Aggregated Preview")
        st.dataframe(summary.head(20))

        # Generate report
        xls_bytes = to_excel_report(summary)

        default_name = f"Yusen_Style_Summary_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button(
            label="Download Excel Report",
            data=xls_bytes,
            file_name=default_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.info(
            "The report includes these tenants even if they have 0 shipments: "
            + ", ".join(REQUIRED_TENANTS)
        )
    except Exception as e:
        st.error(f"Error: {e}")
else:
    st.warning("Please upload the Excel file to proceed.")
